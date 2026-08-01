"""Real cross-process and filesystem tests for the production S4 store.

These tests deliberately use spawned Python interpreters instead of threads or
``multiprocessing``'s platform-dependent default start method.  A passing result
therefore exercises the operating-system lock, handle-release, and sharing
semantics that the retained scheduler will rely on.
"""

import os
from pathlib import Path
import stat
import subprocess
import sys
import threading
import time

from openpyxl import Workbook, load_workbook
import pytest

from harness.evidence import (
    AttemptKey,
    CandidateStateError,
    EvidenceIntegrityError,
    EvidenceStore,
    RunLockedError,
)


RUN_ID = "s4-platform-test"
_DIGESTS = tuple(format(index, "064x") for index in range(1, 8))


_LOCK_HOLDER = r"""
import sys
from pathlib import Path
from harness.evidence import EvidenceStore

store = EvidenceStore.open_run(Path(sys.argv[1]), sys.argv[2])
try:
    with store.locked():
        print("READY", flush=True)
        sys.stdin.buffer.read(1)
except BaseException as exc:
    print(
        "ERROR|{}|{}".format(type(exc).__name__, str(exc)),
        file=sys.stderr,
        flush=True,
    )
    raise
"""


_LOCK_CONTENDER = r"""
import sys
from pathlib import Path
from harness.evidence import EvidenceStore

store = EvidenceStore.open_run(Path(sys.argv[1]), sys.argv[2])
try:
    with store.locked():
        print("ACQUIRED", flush=True)
except BaseException as exc:
    print(
        "BLOCKED|{}|{}".format(type(exc).__name__, str(exc)),
        flush=True,
    )
"""


_WINDOWS_HANDLE_HOLDER = r"""
import ctypes
from ctypes import wintypes
from pathlib import Path
import sys
import time

kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
create_file = kernel32.CreateFileW
create_file.argtypes = (
    wintypes.LPCWSTR,
    wintypes.DWORD,
    wintypes.DWORD,
    wintypes.LPVOID,
    wintypes.DWORD,
    wintypes.DWORD,
    wintypes.HANDLE,
)
create_file.restype = wintypes.HANDLE
close_handle = kernel32.CloseHandle
close_handle.argtypes = (wintypes.HANDLE,)
close_handle.restype = wintypes.BOOL

handle = create_file(
    str(Path(sys.argv[1])),
    0x80000000,  # GENERIC_READ
    0,  # deny read/write/delete sharing
    None,
    3,  # OPEN_EXISTING
    0x80,  # FILE_ATTRIBUTE_NORMAL
    None,
)
if handle == ctypes.c_void_p(-1).value:
    raise ctypes.WinError(ctypes.get_last_error())
try:
    print("READY", flush=True)
    time.sleep(float(sys.argv[2]))
finally:
    if not close_handle(handle):
        raise ctypes.WinError(ctypes.get_last_error())
"""


def _attempt_key(instance_id="platform-instance"):
    """Build the complete, explicit key used by every platform fixture."""

    return AttemptKey(
        domain_name="platform_fixture",
        domain_version="1.0.0",
        domain_content_sha256=_DIGESTS[0],
        task_family="platform_filesystem",
        task_version="1.0.0",
        generator_version="1.0.0",
        grader_version="1.0.0",
        model_tag="platform-model:latest",
        model_digest="sha256:{}".format(_DIGESTS[1]),
        condition_name="harness_full",
        condition_version="1.0.0",
        mechanism_sha256=_DIGESTS[2],
        instance_id=instance_id,
        instance_content_sha256=_DIGESTS[6],
        ordered_subepisodes=("only",),
        repeat=0,
        sampling={"seed": 7, "temperature": "0"},
        opportunity_budget={"model_calls": 1, "tool_calls": 1},
        prompt_sha256=_DIGESTS[3],
        tool_schema_sha256=_DIGESTS[4],
    )


def _run_metadata():
    return {
        "protocol_sha256": _DIGESTS[5],
        "candidate_commit": "1" * 40,
    }


def _create_store(tmp_path, run_id=RUN_ID):
    runs_root = tmp_path / "runs"
    store = EvidenceStore.create_run(runs_root, run_id, _run_metadata())
    return runs_root, store


def _write_valid_evidence(writer):
    """Write the fixed required member set expected by the S4 manifest."""

    writer.write_json(
        "initial-state.json",
        {
            "schema_version": "brick.evidence-state/1",
            "state_kind": "initial",
            "payload": {"counter": 0},
        },
    )
    writer.write_json(
        "final-state.json",
        {
            "schema_version": "brick.evidence-state/1",
            "state_kind": "final",
            "payload": {"counter": 1},
        },
    )
    writer.write_json(
        "result.json",
        {
            "schema_version": "brick.evidence-result/1",
            "execution_status": "done",
            "tool_status": "clean",
            "failure_origin": "none",
            "failure": None,
            "metrics": {},
            "diagnostics": [],
        },
    )
    writer.write_json(
        "grade.json",
        {
            "schema_version": "brick.evidence-grade/1",
            "grader_status": "graded",
            "candidate_decision": True,
            "diagnostics": [],
        },
    )
    writer.write_json(
        "actions.json",
        {"schema_version": "brick.evidence-actions/1", "actions": []},
    )
    writer.write_bytes("transcript.md", b"# Platform fixture\n")
    writer.write_bytes(
        "memory-delta.jsonl",
        b'{"delta":[],"schema_version":"brick.memory-delta/1"}\n',
    )
    workbook_path = writer.artifacts_dir / "probe.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "S4 held-handle fixture"
    sheet.append(("kind", "value"))
    sheet.append(("candidate", writer.path.name))
    workbook.save(workbook_path)
    workbook.close()
    writer.capture_artifact("artifacts/probe.xlsx")


def _physical_candidates(runs_root, run_id=RUN_ID):
    attempts = runs_root / run_id / "attempts"
    if not attempts.is_dir():
        return []
    candidates = []
    for logical in attempts.iterdir():
        if not logical.is_dir():
            continue
        candidates.extend(
            candidate
            for candidate in logical.iterdir()
            if candidate.is_dir()
        )
    return candidates


def _committed_markers(runs_root, run_id=RUN_ID):
    return [
        candidate / "COMMITTED"
        for candidate in _physical_candidates(runs_root, run_id)
        if (candidate / "COMMITTED").exists()
    ]


def _start_lock_holder(runs_root, run_id=RUN_ID):
    process = subprocess.Popen(
        [
            sys.executable,
            "-c",
            _LOCK_HOLDER,
            str(runs_root),
            run_id,
        ],
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    ready = _readline_with_timeout(process, "lock holder")
    if ready != "READY":
        _, stderr = process.communicate(timeout=10)
        raise AssertionError(
            "lock holder did not become ready: {!r}; stderr={!r}".format(
                ready,
                stderr.decode("utf-8", "replace"),
            )
        )
    return process


def _readline_with_timeout(process, label):
    lines = []

    def read_line():
        lines.append(
            process.stdout.readline().decode("utf-8", "replace").strip()
        )

    reader = threading.Thread(target=read_line, daemon=True)
    reader.start()
    reader.join(timeout=10)
    if reader.is_alive():
        process.kill()
        process.communicate(timeout=10)
        reader.join(timeout=1)
        raise AssertionError("{} did not respond within ten seconds".format(label))
    return lines[0] if lines else ""


def _stop_lock_holder(process):
    if process.poll() is None:
        process.stdin.write(b"x")
        process.stdin.flush()
    _, stderr = process.communicate(timeout=10)
    assert process.returncode == 0, stderr.decode("utf-8", "replace")


def _kill_process(process):
    if process.poll() is None:
        process.kill()
    process.communicate(timeout=10)


def _start_windows_handle_holder(path, seconds):
    process = subprocess.Popen(
        [
            sys.executable,
            "-c",
            _WINDOWS_HANDLE_HOLDER,
            str(path),
            str(seconds),
        ],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    ready = _readline_with_timeout(process, "held-handle helper")
    if ready != "READY":
        _, stderr = process.communicate(timeout=10)
        raise AssertionError(
            "held-handle helper failed: {!r}".format(
                stderr.decode("utf-8", "replace")
            )
        )
    return process


def _wait_for_clean_process(process):
    _, stderr = process.communicate(timeout=10)
    assert process.returncode == 0, stderr.decode("utf-8", "replace")


def _lock_error(value):
    lowered = value.lower()
    return "lock" in lowered or "busy" in lowered


def _acquire_after_process_exit(runs_root, timeout_seconds=10.0):
    deadline = time.monotonic() + timeout_seconds
    last_error = None
    while time.monotonic() < deadline:
        store = EvidenceStore.open_run(runs_root, RUN_ID)
        try:
            with store.locked():
                return
        except Exception as exc:
            if not _lock_error(
                "{} {}".format(type(exc).__name__, str(exc))
            ):
                raise
            last_error = exc
            time.sleep(0.05)
    pytest.fail(
        "operating system did not release the killed writer's lock: {!r}".format(
            last_error
        )
    )


def _assert_regular_nonreparse(path):
    info = path.lstat()
    assert stat.S_ISREG(info.st_mode)
    assert not path.is_symlink()
    reparse = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0x400)
    assert not (getattr(info, "st_file_attributes", 0) & reparse)


def _capability_unavailable(reason):
    if os.environ.get("BRICK_S4_NATIVE_REQUIRED") == "1":
        pytest.fail(reason)
    pytest.skip(reason)


def _assert_rejected_without_marker(store, runs_root, key, producer):
    with pytest.raises((EvidenceIntegrityError, CandidateStateError)):
        store.execute_or_resume(key, producer)
    assert _physical_candidates(runs_root)
    assert _committed_markers(runs_root) == []


def test_cross_process_lock_excludes_contender_and_persists(tmp_path):
    runs_root, _ = _create_store(tmp_path)
    lock_path = runs_root / RUN_ID / "run.lock"
    _assert_regular_nonreparse(lock_path)
    identity_before = lock_path.stat()

    holder = _start_lock_holder(runs_root)
    try:
        contender = subprocess.run(
            [
                sys.executable,
                "-c",
                _LOCK_CONTENDER,
                str(runs_root),
                RUN_ID,
            ],
            check=True,
            capture_output=True,
            text=True,
            timeout=10,
        )
        output = contender.stdout.strip()
        assert output.startswith("BLOCKED|"), output
        assert _lock_error(output), output
        _assert_regular_nonreparse(lock_path)
        identity_during = lock_path.stat()
        if identity_before.st_ino and identity_during.st_ino:
            assert (identity_before.st_dev, identity_before.st_ino) == (
                identity_during.st_dev,
                identity_during.st_ino,
            )
    finally:
        _stop_lock_holder(holder)

    with EvidenceStore.open_run(runs_root, RUN_ID).locked():
        _assert_regular_nonreparse(lock_path)
    identity_after = lock_path.stat()
    if identity_before.st_ino and identity_after.st_ino:
        assert (identity_before.st_dev, identity_before.st_ino) == (
            identity_after.st_dev,
            identity_after.st_ino,
        )


def test_cross_process_lock_is_released_after_forced_termination(tmp_path):
    runs_root, _ = _create_store(tmp_path)
    lock_path = runs_root / RUN_ID / "run.lock"
    identity_before = lock_path.stat()

    holder = _start_lock_holder(runs_root)
    _kill_process(holder)
    _acquire_after_process_exit(runs_root)

    _assert_regular_nonreparse(lock_path)
    identity_after = lock_path.stat()
    if identity_before.st_ino and identity_after.st_ino:
        assert (identity_before.st_dev, identity_before.st_ino) == (
            identity_after.st_dev,
            identity_after.st_ino,
        )


def test_execute_acquires_lock_before_candidate_or_producer(tmp_path):
    runs_root, store = _create_store(tmp_path)
    called = []

    def producer(writer):
        called.append(writer.path)
        _write_valid_evidence(writer)

    holder = _start_lock_holder(runs_root)
    try:
        with pytest.raises(RunLockedError):
            store.execute_or_resume(_attempt_key(), producer)
    finally:
        _stop_lock_holder(holder)

    assert called == []
    assert _physical_candidates(runs_root) == []


@pytest.mark.skipif(os.name == "nt", reason="POSIX filesystem fixture")
def test_posix_symlink_member_is_rejected_before_publication(tmp_path):
    runs_root, store = _create_store(tmp_path)

    def producer(writer):
        _write_valid_evidence(writer)
        (writer.path / "artifacts" / "escape").symlink_to(
            writer.path / "initial-state.json"
        )

    _assert_rejected_without_marker(
        store,
        runs_root,
        _attempt_key("posix-symlink"),
        producer,
    )


@pytest.mark.skipif(os.name == "nt", reason="POSIX filesystem fixture")
def test_posix_irregular_member_is_rejected_before_publication(tmp_path):
    if not hasattr(os, "mkfifo"):
        pytest.skip("os.mkfifo is unavailable")
    runs_root, store = _create_store(tmp_path)

    def producer(writer):
        _write_valid_evidence(writer)
        os.mkfifo(str(writer.path / "artifacts" / "irregular"))

    _assert_rejected_without_marker(
        store,
        runs_root,
        _attempt_key("posix-fifo"),
        producer,
    )


@pytest.mark.skipif(os.name != "nt", reason="Windows reparse fixture")
@pytest.mark.parametrize("target_is_directory", [False, True])
def test_windows_real_symlink_is_rejected_before_publication(
    tmp_path,
    target_is_directory,
):
    probe_target = tmp_path / "symlink-probe-target"
    probe_link = tmp_path / "symlink-probe-link"
    if target_is_directory:
        probe_target.mkdir()
    else:
        probe_target.write_bytes(b"probe")
    try:
        probe_link.symlink_to(
            probe_target,
            target_is_directory=target_is_directory,
        )
    except OSError as exc:
        _capability_unavailable(
            "Windows symlink creation is unavailable; enable Developer Mode: "
            "{!r}".format(exc)
        )
    else:
        if target_is_directory:
            probe_link.rmdir()
        else:
            probe_link.unlink()

    runs_root, store = _create_store(tmp_path)

    def producer(writer):
        _write_valid_evidence(writer)
        link = writer.path / "artifacts" / "reparse-link"
        target = (
            writer.path / "artifacts"
            if target_is_directory
            else writer.path / "initial-state.json"
        )
        link.symlink_to(target, target_is_directory=target_is_directory)

    try:
        _assert_rejected_without_marker(
            store,
            runs_root,
            _attempt_key(
                "windows-directory-symlink"
                if target_is_directory
                else "windows-file-symlink"
            ),
            producer,
        )
    finally:
        for candidate in _physical_candidates(runs_root):
            link = candidate / "artifacts" / "reparse-link"
            if link.is_symlink():
                if target_is_directory:
                    link.rmdir()
                else:
                    link.unlink()


def _create_windows_junction(link, target):
    completed = subprocess.run(
        [
            "cmd.exe",
            "/d",
            "/c",
            "mklink",
            "/J",
            str(link),
            str(target),
        ],
        capture_output=True,
        text=True,
        timeout=10,
    )
    if completed.returncode != 0:
        raise OSError(
            "mklink /J failed: {}".format(
                (completed.stderr or completed.stdout).strip()
            )
        )


@pytest.mark.skipif(os.name != "nt", reason="Windows junction fixture")
def test_windows_real_junction_is_rejected_before_publication(tmp_path):
    external = tmp_path / "junction-target"
    external.mkdir()
    probe = tmp_path / "junction-probe"
    try:
        _create_windows_junction(probe, external)
    except OSError as exc:
        _capability_unavailable(
            "Windows junction creation is unavailable: {!r}".format(exc)
        )
    else:
        probe.rmdir()

    runs_root, store = _create_store(tmp_path)

    def producer(writer):
        _write_valid_evidence(writer)
        _create_windows_junction(
            writer.path / "artifacts" / "junction",
            external,
        )

    try:
        _assert_rejected_without_marker(
            store,
            runs_root,
            _attempt_key("windows-junction"),
            producer,
        )
    finally:
        for candidate in _physical_candidates(runs_root):
            junction = candidate / "artifacts" / "junction"
            if junction.exists():
                junction.rmdir()


@pytest.mark.skipif(os.name != "nt", reason="Windows held-handle fixture")
def test_windows_real_held_handle_retries_without_model_rerun(tmp_path):
    runs_root, store = _create_store(tmp_path)
    producer_calls = []

    def producer(writer):
        producer_calls.append(1)
        _write_valid_evidence(writer)

    initial = store.execute_or_resume(
        _attempt_key("windows-held-handle"),
        producer,
        deadline_seconds=5,
    )
    assert initial.state == "committed"
    candidate = initial.candidate_path
    reopened = load_workbook(
        candidate / "artifacts" / "probe.xlsx",
        read_only=True,
        data_only=True,
    )
    try:
        assert reopened["S4 held-handle fixture"]["A2"].value == "candidate"
    finally:
        reopened.close()
    (candidate / "COMMITTED").unlink()
    holder = _start_windows_handle_holder(
        candidate / "artifacts" / "probe.xlsx",
        0.5,
    )

    started = time.monotonic()
    try:
        result = store.execute_or_resume(
            _attempt_key("windows-held-handle"),
            producer,
            deadline_seconds=5,
        )
    finally:
        _wait_for_clean_process(holder)

    assert producer_calls == [1]
    assert result.state == "committed"
    assert result.producer_called is False
    assert time.monotonic() - started >= 0.2
    markers = _committed_markers(runs_root)
    assert len(markers) == 1
    assert markers[0].read_bytes() == b""


@pytest.mark.skipif(os.name != "nt", reason="Windows held-handle fixture")
def test_windows_held_handle_timeout_recovers_without_model_rerun(tmp_path):
    runs_root, store = _create_store(tmp_path)
    producer_calls = []

    def producer(writer):
        producer_calls.append(1)
        _write_valid_evidence(writer)

    initial = store.execute_or_resume(
        _attempt_key("windows-held-handle-recovery"),
        producer,
        deadline_seconds=5,
    )
    assert initial.state == "committed"
    candidate = initial.candidate_path
    reopened = load_workbook(
        candidate / "artifacts" / "probe.xlsx",
        read_only=True,
        data_only=True,
    )
    try:
        assert reopened["S4 held-handle fixture"]["A2"].value == "candidate"
    finally:
        reopened.close()
    (candidate / "COMMITTED").unlink()
    holder = _start_windows_handle_holder(
        candidate / "artifacts" / "probe.xlsx",
        0.5,
    )

    try:
        first = store.execute_or_resume(
            _attempt_key("windows-held-handle-recovery"),
            producer,
            deadline_seconds=0.05,
        )
    finally:
        _wait_for_clean_process(holder)
    assert first.state == "publish_blocked"
    assert first.producer_called is False
    assert producer_calls == [1]
    assert _committed_markers(runs_root) == []

    resumed = store.execute_or_resume(
        _attempt_key("windows-held-handle-recovery"),
        producer,
        deadline_seconds=5,
    )
    assert resumed.state == "committed"
    assert resumed.producer_called is False
    assert producer_calls == [1]
    markers = _committed_markers(runs_root)
    assert len(markers) == 1
    assert markers[0].read_bytes() == b""
